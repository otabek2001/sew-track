"""
Reports and Export utilities for Admin Panel.
"""

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from datetime import date, datetime, timedelta
from django.db.models import Sum

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from apps.tasks.models import WorkRecord
from apps.employees.models import Employee


def export_daily_report_excel(tenant, report_date):
    """Export daily work records report to Excel."""
    
    # Get records for the day
    records = WorkRecord.objects.filter(
        tenant=tenant,
        work_date=report_date
    ).select_related(
        'employee', 'product', 'task', 'product_task'
    ).order_by('employee__full_name', 'created_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Kunlik hisobot"
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"{tenant.name} - Kunlik hisobot"
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.row_dimensions[1].height = 30
    
    # Date
    ws.merge_cells('A2:H2')
    date_cell = ws['A2']
    date_cell.value = f"Sana: {report_date.strftime('%d.%m.%Y')}"
    date_cell.font = Font(size=12, bold=True)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    ws.append([])  # Empty row
    headers = ['#', 'Xodim', 'Mahsulot', 'Operatsiya', 'Miqdor', 'Narx', 'Jami', 'Status']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 9):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    row_num = 5
    total_quantity = 0
    total_payment = 0
    
    status_map = {
        'pending': 'Kutilmoqda',
        'approved': 'Tasdiqlangan',
        'rejected': 'Rad etilgan',
        'completed': 'Bajarilgan'
    }
    
    for idx, record in enumerate(records, 1):
        row_data = [
            idx,
            record.employee.full_name,
            record.product.name if record.product else '-',
            record.task.name_uz if record.task else '-',
            record.quantity,
            record.unit_price,
            record.total_payment,
            status_map.get(record.status, record.status)
        ]
        
        ws.append(row_data)
        
        # Apply border to all cells in row
        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col in [5, 6, 7]:  # Numbers
                cell.alignment = Alignment(horizontal='right')
        
        total_quantity += record.quantity
        total_payment += record.total_payment
        row_num += 1
    
    # Totals row
    ws.append([])
    totals_row = ['', '', '', 'JAMI:', total_quantity, '', total_payment, '']
    ws.append(totals_row)
    
    # Style totals
    for col in range(1, 9):
        cell = ws.cell(row=row_num + 1, column=col)
        cell.font = Font(bold=True)
        cell.border = border
        if col in [5, 7]:
            cell.alignment = Alignment(horizontal='right')
    
    # Adjust column widths
    column_widths = [5, 25, 20, 20, 10, 12, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"kunlik_hisobot_{report_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def export_employee_report_excel(employee, start_date, end_date):
    """Export individual employee report to Excel."""
    
    # Get records
    records = WorkRecord.objects.filter(
        employee=employee,
        work_date__gte=start_date,
        work_date__lte=end_date
    ).select_related(
        'product', 'task', 'product_task'
    ).order_by('work_date', 'created_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Xodim hisoboti"
    
    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"{employee.full_name} - Shaxsiy hisobot"
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.row_dimensions[1].height = 30
    
    # Period
    ws.merge_cells('A2:G2')
    period_cell = ws['A2']
    period_cell.value = f"Davr: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    period_cell.font = Font(size=12, bold=True)
    period_cell.alignment = Alignment(horizontal='center')
    
    # Summary stats
    ws.append([])
    total_records = records.count()
    total_quantity = records.aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_payment = records.aggregate(Sum('total_payment'))['total_payment__sum'] or 0
    approved_records = records.filter(status='approved').count()
    
    ws.append(['Jami yozuvlar:', total_records])
    ws.append(['Jami mahsulot:', total_quantity])
    ws.append(['Jami to\'lov:', f"{total_payment:,.0f} so'm"])
    ws.append(['Tasdiqlangan:', approved_records])
    
    # Headers
    ws.append([])
    headers = ['Sana', 'Mahsulot', 'Operatsiya', 'Miqdor', 'Narx', 'Jami', 'Status']
    ws.append(headers)
    
    # Style headers
    header_row = ws.max_row
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 8):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    status_map = {
        'pending': 'Kutilmoqda',
        'approved': 'Tasdiqlangan',
        'rejected': 'Rad etilgan',
        'completed': 'Bajarilgan'
    }
    
    for record in records:
        row_data = [
            record.work_date.strftime('%d.%m.%Y'),
            record.product.name if record.product else '-',
            record.task.name_uz if record.task else '-',
            record.quantity,
            record.unit_price,
            record.total_payment,
            status_map.get(record.status, record.status)
        ]
        
        ws.append(row_data)
        
        # Apply border
        row_num = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col in [4, 5, 6]:  # Numbers
                cell.alignment = Alignment(horizontal='right')
    
    # Adjust column widths
    column_widths = [12, 20, 20, 10, 12, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{employee.full_name}_hisobot_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def export_monthly_summary_excel(tenant, year, month):
    """Export monthly summary report (all employees)."""
    
    # Get first and last day of month
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # Get all employees
    employees = Employee.objects.filter(
        tenant=tenant,
        is_active=True
    ).order_by('full_name')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Oylik hisobot"
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    month_names = ['', 'Yanvar', 'Fevral', 'Mart', 'Aprel', 'May', 'Iyun',
                   'Iyul', 'Avgust', 'Sentabr', 'Oktabr', 'Noyabr', 'Dekabr']
    title_cell.value = f"{tenant.name} - {month_names[month]} {year} oylik hisobot"
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.row_dimensions[1].height = 30
    
    # Headers
    ws.append([])
    headers = ['#', 'Xodim', 'Yozuvlar soni', 'Jami mahsulot', 'Tasdiqlangan', 'Jami to\'lov (so\'m)']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 7):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    row_num = 4
    grand_total_records = 0
    grand_total_quantity = 0
    grand_total_approved = 0
    grand_total_payment = 0
    
    for idx, employee in enumerate(employees, 1):
        records = WorkRecord.objects.filter(
            employee=employee,
            work_date__gte=first_day,
            work_date__lte=last_day
        )
        
        total_records = records.count()
        total_quantity = records.aggregate(Sum('quantity'))['quantity__sum'] or 0
        approved_count = records.filter(status='approved').count()
        total_payment = records.filter(status='approved').aggregate(
            Sum('total_payment')
        )['total_payment__sum'] or 0
        
        row_data = [
            idx,
            employee.full_name,
            total_records,
            total_quantity,
            approved_count,
            total_payment
        ]
        
        ws.append(row_data)
        
        # Apply border
        for col in range(1, 7):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col in [3, 4, 5, 6]:  # Numbers
                cell.alignment = Alignment(horizontal='right')
        
        grand_total_records += total_records
        grand_total_quantity += total_quantity
        grand_total_approved += approved_count
        grand_total_payment += total_payment
        row_num += 1
    
    # Totals row
    ws.append([])
    totals_row = ['', 'JAMI:', grand_total_records, grand_total_quantity, grand_total_approved, grand_total_payment]
    ws.append(totals_row)
    
    # Style totals
    for col in range(1, 7):
        cell = ws.cell(row=row_num + 1, column=col)
        cell.font = Font(bold=True, size=12)
        cell.border = border
        if col in [3, 4, 5, 6]:
            cell.alignment = Alignment(horizontal='right')
        if col == 6:
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # Adjust column widths
    column_widths = [5, 25, 15, 15, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"oylik_hisobot_{year}_{month:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

